from __future__ import division
import cv2
from statistics import mean 
import skimage.exposure as sk_exposure
import numpy as np
import os
import statistics
import matplotlib.pyplot as plt
from skimage.measure import find_contours
from skimage import img_as_ubyte
from scipy.spatial import Delaunay
from histomicstk.preprocessing.color_normalization import reinhard
from histomicstk.preprocessing.color_conversion import lab_mean_std
import random as rng
import math
import alphashape
from skimage import color
from statistics import mean
from histmatch import hist_match
from skimage.color import separate_stains,fgx_from_rgb
from matplotlib.colors import LinearSegmentedColormap
from skimage.color import rgb2hed
from colortransfert import color_transfer
from lumen_seg_2 import *
from scipy.signal import find_peaks
from skimage.morphology import label
from skimage.io import imshow
import imutils
from scipy import ndimage
from itertools import combinations
import scipy as sp
import skimage.io
import matplotlib.pyplot as plt
from skimage.morphology import label
from skimage.morphology import watershed
from scipy import asarray as ar,exp
from scipy.spatial import Delaunay
from skimage.color import rgb2hed
from skimage.segmentation import slic
from skimage.segmentation import mark_boundaries
from statistics import stdev 
from openpyxl import Workbook
from openpyxl import load_workbook
from skimage import morphology
from skimage.feature import greycomatrix, greycoprops
from skimage.segmentation import felzenszwalb
import warnings
warnings.filterwarnings('ignore')
def get_mpl_colormap(cmap):
    

    # Initialize the matplotlib color map
    sm = plt.cm.ScalarMappable(cmap=cmap)

    # Obtain linear color range
    color_range = sm.to_rgba(np.linspace(0, 1, 256), bytes=True)[:,2::-1]

    return color_range.reshape(256, 1, 3)

def get_features_conf(indice_folder):
    file_excel="C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/output/Norm_3.xlsx"
    source_image= cv2.imread('C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/codes/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png',-1)
    meanRef, stdRef = lab_mean_std(source_image)

    list_img_final_label = []
    wb=load_workbook(file_excel)
    sheet = wb.active

   
    suffix = str(indice_folder).zfill(3)
    path_mask_Glands_= 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_glands'
    path_mask_Nuclei_rgb_ = 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/output_rgb'
    path_image_RGB_ = 'C:/Users/Laura/AppData/Local/Programs/Python/Python36/Phenotype/Unet_nuclei/wsi/tiles_png'
    path_final = 'C:/Users/laura/AppData/Local/Programs/Python/Python36/Phenotype/Output/Output_final'
#for num, name_files in (enumerate(os.listdir(image_predic_path))):
 #   print(name_files)/
#name_image_mask = path_mask_Glands +  '/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png '# TCGA-001-tile-r6-c110-x111622-y5122-w1024-h1024_class #_class/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png'
#name_image = path_image_RGB + '/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png '#'/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024.png'
#image_nuclei =path_mask_Nuclei+'/TCGA-001-tile-r5-c109-x110598-y4098-w1024-h1024_class.png'


    indice_row = 2
    OUTPUT_DIR = path_final + '/' +suffix
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

   
    
    str1 = ''.join(suffix)
    sheet.cell(row=indice_row, column=2).value=str(str1)
  
    path_mask_Glands = path_mask_Glands_ +'/'+suffix
  #  path_mask_Nuclei_hed=path_mask_Nuclei_hed_  +'/'+suffix
    path_mask_Nuclei_rgb=path_mask_Nuclei_rgb_  +'/'+suffix
    path_image_RGB = path_image_RGB_ +'/'+suffix

    list_glands =os.listdir(path_mask_Glands)
    list_glands = set(list_glands)
  #  list_nuclei_hed =os.listdir(path_mask_Nuclei_hed)
  #  list_nuclei_hed = set(list_nuclei_hed)
    list_nuclei_rgb =os.listdir(path_mask_Nuclei_rgb)
    list_nuclei_rgb = set(list_nuclei_rgb)
    Ratio_Glands_lumen = []
   
    
 

    Distance_between_glands = []
    disper_type = []
    for num, fname in (enumerate(os.listdir(path_image_RGB  ))):
        
        width = 364
        height = 364
        dim = (width, height)
        Final_mask_full = np.zeros((width, height), dtype=np.uint8)
                  

        
        indice_fused = 0
        indice_indivi =0
        indice_glands_dis = 0
        indice_glands_without_lumen = 0
  
        indice_nber_glands = 0



        list_ratio_non_lumen_to_contour = []
        list_ratio_Glands_lumen = []
        list_ratio_contour_2= []
        list_ratio_ROI = []
        list_ratio_contour = []
        list_ratio_density_stroma = []
        list_ratio_density_nuclei= []
        list_value_mac = []
        list_Variance_shape_glands = []
        list_space_glands = []
        list_lumen_true=[]
        list_ratio_density_nuclei_1=[]

                                      
        p= os.path.basename(fname)
        name1 = os.path.splitext(p)[0]
        
       
      #  wb.save(file_excel)
     #   matching_nuclei_hed = [s for s in list_nuclei_hed  if name1 in s]
        matching_glands =[g for g in list_glands  if name1 in g]
        matching_nuclei_rgb = [h for h in list_nuclei_rgb  if name1 in h]
        
        if len(matching_nuclei_rgb)==0 or len(matching_glands)==0    :
            continue
        elif len(matching_nuclei_rgb)!= 0 and len(matching_glands)!= 0  :
            
            width = 364
            height = 364
            dim = (width, height)
        
            name_image = path_image_RGB +'/'+  fname
            img_glands_RGB_ = cv2.imread(name_image ,-1)
            img_glands_RGB= cv2.resize(img_glands_RGB_, dim, interpolation = cv2.INTER_AREA)
            img_glands_RGB_reinhard = reinhard (img_glands_RGB, meanRef, stdRef)
            img_hsv = cv2.cvtColor(img_glands_RGB_reinhard, cv2.COLOR_BGR2HSV)
            blank_image = np.zeros((height,width,3), np.uint8)
            blank_image[:]=(139,58,98)


            final_lumen_mask = np.zeros((height,width), np.uint8)
          
          
            
            str3 = ''.join(matching_nuclei_rgb)
            image_nuclei_rgb = path_mask_Nuclei_rgb +'/'+ str3
            img_nuclei_mask_rgb = cv2.imread(image_nuclei_rgb,0)
            img_nuclei_mask_rgb= cv2.resize(img_nuclei_mask_rgb, dim, interpolation = cv2.INTER_AREA)

         #   str4 = ''.join(matching_nuclei_hed)
         #   image_nuclei_hed = path_mask_Nuclei_hed +'/'+  str4
         #   img_nuclei_mask_hed= cv2.imread(image_nuclei_hed,0)
         #   img_nuclei_mask_hed= cv2.resize(img_nuclei_mask_hed, dim, interpolation = cv2.INTER_AREA)

####### Glands ROI ######
            str2 = ''.join(matching_glands)
            name_image_mask = path_mask_Glands +'/'+ str2
            img_glands_mask = cv2.imread(name_image_mask,0)
            img_glands_mask = zoom(img_glands_mask, 2)
            img_glands_mask  = cv2.resize(img_glands_mask , dim, interpolation = cv2.INTER_AREA)

            centerLumen_X_bitwise= []
            centerLumen_Y_bitwise = []
            Contour_lumen_bitwise= []
            
            img_nuclei_mask = img_nuclei_mask_rgb #+ img_nuclei_mask_hed
            
            shape = np.shape(img_glands_RGB)
            Mask_image_rgb = np.zeros((shape[0], shape[1],3), dtype=np.uint8)
            
            Final_image_lum = np.zeros((shape[0], shape[1]), dtype=np.uint8)


            mask_rec = cv2.bitwise_not(img_glands_mask )
            try2_2 = cv2.bitwise_and(img_glands_RGB,img_glands_RGB, mask=img_glands_mask)
            img_glands_RGB_inside = cv2.cvtColor(img_glands_RGB,cv2.COLOR_BGR2RGB)
            img_inside = cv2.cvtColor(mask_rec,cv2.COLOR_BGR2RGB)
            _, mask2_inside = cv2.threshold(mask_rec,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            contoursMask_inside, hierarchy_inside = cv2.findContours(mask2_inside,cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            kernel_dil = np.ones((2,2), np.uint8)
            ###keep lumen
            Img_final,centerLumen_X_bitwise,centerLumen_Y_bitwise,Contour_lumen_bitwise = Img_reslt_fin(Contour_lumen_bitwise,try2_2,Mask_image_rgb,contoursMask_inside,hierarchy_inside,img_glands_RGB,img_glands_RGB_reinhard,centerLumen_X_bitwise,centerLumen_Y_bitwise)
            ##### delete stroma
           
            img_slic,mask_good ,img_deconv = image_filter_equa(Img_final,img_glands_RGB_reinhard) 
            segments = slic(img_deconv,compactness=0.1, n_segments =2900,sigma = 0.5)
            superpixels = color.label2rgb(segments,img_slic , kind='avg')


            img_hsv_ = cv2.cvtColor(Img_final, cv2.COLOR_BGR2HSV)

            cv2.imshow('',img_hsv_)
            cv2.waitKey(0)
            
          #  img_glands_RGB = cv2.cvtColor(img_glands_RGB,cv2.COLOR_BGR2RGB)
            img = cv2.cvtColor(img_glands_mask,cv2.COLOR_BGR2RGB)
            _, mask2 = cv2.threshold(img_glands_mask,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            res = cv2.bitwise_and(img,img, mask=mask2)
            
            hsv2bgr = cv2.cvtColor(res, cv2.COLOR_HSV2BGR)
            rgb2gray = cv2.cvtColor(hsv2bgr, cv2.COLOR_BGR2GRAY)
            contoursMask, hierarchy = cv2.findContours(rgb2gray,cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            contoursMask, hierarchy = cv2.findContours(rgb2gray,cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            cv2.drawContours(img_glands_RGB ,contoursMask, -1, (0,255,0), 1)
            cv2.imshow('',img_glands_RGB)
            cv2.waitKey(0)
            

####LUMEN####

            cmap_eosin = LinearSegmentedColormap.from_list('mycmap', ['darkviolet',
                                               'white'])


#
            area_thr_mask=500
            if len(contoursMask) == 0:
                break
            else:
                for c in zip(contoursMask, hierarchy[0]):
                    indice_glands_dis_lumen = 0
                    Final_image_mask_added = np.zeros((shape[0], shape[1]), dtype=np.uint8)
                    
                    height_full_image, width_full_image= img_glands_mask.shape 
                    area_full_image = height_full_image * width_full_image 
  
                    if cv2.contourArea(c[0]) > area_thr_mask and c[1][3] == -1:
                        
        ########LUMEN#####
        
                        centerLumen_X= []
                        centerLumen_Y = []
                        area_lumen=[]
                        perimeter_lumen=[]
                        ellipse = []
                        Contour_lumen = []
                        corner_tab_0 = []
                        corner_tab_1= []
                        indice_lumen=0
                      
                    
                        contour_glands_final_no_lumen = cv2.contourArea(c[0])
                        if c[1][3] == -1 :
                            #c[1][2] != -1
                            temp = np.zeros(img_glands_RGB.shape, dtype=np.uint8)
                            cv2.fillPoly(temp, pts=[c[0]], color=(255, 255, 255))

                            area_lumen,masked_image_nuclei_,Contour_lumen, masked_imageGlands_Black,imgResult,centerLumen_X,centerLumen_Y,indice_lumen,corner_tab_0,corner_tab_1=convex_glands(Contour_lumen_bitwise,centerLumen_X_bitwise,centerLumen_Y_bitwise,kernel_dil,Img_final ,img_nuclei_mask,c,img_glands_RGB,img_glands_RGB_reinhard,centerLumen_X,centerLumen_Y,indice_lumen,area_lumen,Contour_lumen,corner_tab_0,corner_tab_1)
      ######
                
                        
                        Nuclei_mask_separated=  masked_image_nuclei_[0]
                      #  img_erosion_nuclei= cv2.erode(Nuclei_mask_separated, kernel_dil, iterations=2)
                  
                        img_dilation_nuclei= cv2.dilate(Nuclei_mask_separated, kernel_dil, iterations=5)
                       
                        ph_0 = np.ones((img_glands_RGB.shape[0], img_glands_RGB.shape[1], 3), dtype='uint8')
                        ph_0[:,:,0] = img_dilation_nuclei
                        ph_0[:,:,1] = img_dilation_nuclei
                        ph_0[:,:,2] = img_dilation_nuclei
                        nuclei_only = cv2.bitwise_and(blank_image, ph_0)
                        img_erosion_nuclei= cv2.erode(nuclei_only, kernel_dil, iterations=2)
                        
                       
                       
                       #  
                        Nuclei_minus_glands = temp[:, :, 1] -Nuclei_mask_separated
                        if indice_lumen == 0  :
                            indice_glands_without_lumen = indice_glands_without_lumen +1
                              
                            ratio_contour =  (contour_glands_final_no_lumen /area_full_image)*100
                            list_ratio_contour.append(ratio_contour)
                            indices_mask_nucl= np.where(Nuclei_mask_separated == [255])
                            indices_mask_temp= np.where(temp == [255])
                            ratio_density_nuclei_2 = len(indices_mask_nucl[0])/len(indices_mask_temp[0])
                            list_ratio_density_nuclei.append(ratio_density_nuclei_2)
                        if indice_lumen > 0 :
                             final_lumen_mask =final_lumen_mask + masked_imageGlands_Black[:, :, 1]
                             if indice_lumen  == 1 :
                                indice_indivi=indice_indivi+1
                             else :
                                indice_fused = indice_fused + 1
                             #Nuclei_minus_glands =  temp[:, :, 1] +masked_imageGlands_Black[:, :, 1]
                            
                             Nuclei_minus_glands = Nuclei_minus_glands-masked_imageGlands_Black[:, :, 1]

                             
                             indices_bitwiseOr_convex = np.where(Nuclei_mask_separated == [255])
                             density_bit = len(indices_bitwiseOr_convex[0])
   
                             indices_mask_convex = np.where(temp == [255])
                             density_convex = len(indices_mask_convex[0])
                             ratio_density_nuclei = density_bit /density_convex

                             list_ratio_density_nuclei.append(ratio_density_nuclei)
                             
                             ph = np.ones((Nuclei_minus_glands.shape[0], Nuclei_minus_glands.shape[1], 3), dtype='uint8')
                             ph[:,:,0] = Nuclei_minus_glands
                             ph[:,:,1] = Nuclei_minus_glands
                             ph[:,:,2] = Nuclei_minus_glands

                           
                             masked_image_ = cv2.bitwise_and(img_hsv, ph)
                             masked_image_ll = cv2.bitwise_and(img_glands_RGB, temp)
                             equa_dd = equalize_adapthist(masked_image_)
                             saturation = equa_dd [:,:,1] < 0.35
                             saturation_2 = equa_dd [:,:,1] > 0.1
                             
                             mask = saturation*saturation_2
                             
                             Nuclei_minus_glands_2 = Nuclei_minus_glands*mask
                            
                             ph_2 = np.ones((Nuclei_minus_glands.shape[0], Nuclei_minus_glands.shape[1], 3), dtype='uint8')
                             ph_2[:,:,0] = Nuclei_minus_glands_2
                             ph_2[:,:,1] = Nuclei_minus_glands_2
                             ph_2[:,:,2] = Nuclei_minus_glands_2
                             bitwiseOr = cv2.bitwise_or(Nuclei_minus_glands_2, Nuclei_mask_separated)
                            
                             bitwiseOr_with_glands = cv2.bitwise_or(bitwiseOr, masked_imageGlands_Black[:,:,1])
                             #bitwiseOr_with_glands = cv2.bitwise_or(Nuclei_minus_glands_2, masked_imageGlands_Black[:,:,1])
                             
                            
                            # img_dilation_cleaned = cv2.erode(cleaned_, kernel_dil, iterations=2)
                  
                             
                             bitwiseOr_2 = np.ones((Nuclei_minus_glands.shape[0], Nuclei_minus_glands.shape[1], 3), dtype='uint8')
                             bitwiseOr_2[:,:,0] = bitwiseOr_with_glands
                             bitwiseOr_2[:,:,1] = bitwiseOr_with_glands
                             bitwiseOr_2[:,:,2] = bitwiseOr_with_glands
                             masked_image_ll_2 = cv2.bitwise_and(img_hsv, bitwiseOr_2)
                            # cv2.imshow('',bitwiseOr_with_glands)
                            # cv2.waitKey(0)

                             mask_with_belt_nuclei = masked_image_ll_2 +img_erosion_nuclei
                          
                            
                             

                             masked_image_ll_2_equa = equalize_adapthist(masked_image_ll_2)
                             img_dilation_cleaned = cv2.dilate(masked_image_ll_2_equa, kernel_dil, iterations=2)
                             
                             dist = cv2.distanceTransform(bitwiseOr_with_glands, cv2.DIST_L2, 3)
                             
                             cv2.normalize(dist, dist, 0, 1.0, cv2.NORM_MINMAX)
                             
                             _, dist = cv2.threshold(dist, 0.45, 1.0, cv2.THRESH_BINARY)
                             
                             dist = dist.astype('uint8')
                             _, mask2_dist= cv2.threshold(dist  ,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                             contours_dist, hierarchy_dist = cv2.findContours(mask2_dist,cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)

                             lum = masked_imageGlands_Black[:, :, 1].astype('uint8')
                            
                             _, lumen_per_contour= cv2.threshold(lum  ,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                             contours_lumen, hierarchy_lum = cv2.findContours(lumen_per_contour,cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
                             cv2.drawContours(img_glands_RGB ,contours_dist, -1, (0,255,0), 1)
                             cv2.imshow('',img_glands_RGB)
                             cv2.waitKey(0)
                             Final_image_mask_added_existence = False
                             for c_lumen in zip(contours_lumen, hierarchy_lum[0]):
                                Lumen_dis_ok =  False
                                temp_lumen_dist = np.zeros(img_glands_RGB.shape, dtype=np.uint8)
                                cv2.fillPoly(temp_lumen_dist, pts=[c_lumen[0]], color=(255, 255, 255))
                                area_lum  = cv2.contourArea(c_lumen[0])
                                if cv2.contourArea(c_lumen[0])>10:
                                    for c_dist in zip(contours_dist, hierarchy_dist[0]):
                                        if cv2.contourArea(c_dist[0])>2:
                                            M = cv2.moments(c_dist[0])
                                            cX_lumen__= int(M["m10"] / M["m00"])
                                            cY_lumen__ = int(M["m01"] / M["m00"])
                                            temp__dist = np.zeros(img_glands_RGB.shape, dtype=np.uint8)
                                            cv2.fillPoly(temp__dist, pts=[c_dist[0]], color=(255, 255, 255))  
                                            overlay_ = cv2.pointPolygonTest(c_lumen[0],(cX_lumen__,cY_lumen__),True)
                                            if overlay_> 0 :
                                                bitwiseXor = cv2.bitwise_or(temp_lumen_dist, temp__dist)
                                                _, mask_add= cv2.threshold(bitwiseXor[:, :, 1]  ,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                                                contours_add, hierarchy_add = cv2.findContours(mask_add,cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
                                                for c_add in zip(contours_add, hierarchy_add[0]):
                                                    area_add  = cv2.contourArea(c_add[0])
                                                    if area_lum /area_add>0.98 and area_lum /area_add<1.1 :
                                                        Lumen_dis_ok =  True
                                                    else:
                                                        Lumen_dis_ok =  False
                                                        
                                if  Lumen_dis_ok == False :
                                    Final_image_mask_added_existence =  True
                                    
                                    Final_image_mask_added = Final_image_mask_added + temp_lumen_dist[:, :, 1]
                                    Final_mask_full = Final_mask_full +temp_lumen_dist[:, :, 1]
                                    indice_glands_dis = indice_glands_dis+1
                                    indice_glands_dis_lumen = indice_glands_dis_lumen+1
                                    ratio_non_lumen_to_contour =area_lum/contour_glands_final_no_lumen
                                    list_ratio_non_lumen_to_contour.append(ratio_non_lumen_to_contour)
                                    indice_nber_glands_true = indice_glands_dis/indice_lumen
                                    list_lumen_true.append(indice_nber_glands_true)  
                                if  Lumen_dis_ok == True :              
                                    Final_lumen_mask_added = False
                        
                                    evolution = []
                                    callback = store_evolution_in(evolution)
                                    imgResult23 = img_as_float64(masked_image_ll_2_equa)
                                 
                                    ls = morphological_geodesic_active_contour(imgResult23, 250, temp_lumen_dist, threshold=0.5,balloon = 1,
                                           smoothing=1,iter_callback=callback) #0.20
                                    
                                    ls = ls.astype('uint8')
                                    ls = 255*ls
                                    _, mask_ls= cv2.threshold(ls[:, :, 1]  ,0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                                      
                                    contours_ls, hierarchy_add = cv2.findContours(mask_ls,cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
                                    cmax_ = max(contours_ls, key = cv2.contourArea)
                                    
                                    contour_glands_final = cv2.contourArea(cmax_)
                                    ratio_Glands_lumen = (area_lum/contour_glands_final)*100
                                    list_ratio_Glands_lumen.append(ratio_Glands_lumen)
                                    
                                    ratio_contour_2 = (contour_glands_final /area_full_image)*100
                                    list_ratio_contour_2.append(ratio_contour_2)
                                   
                                    ratio_ROI = (contour_glands_final/contour_glands_final_no_lumen)*100
                                    list_ratio_ROI .append(ratio_ROI)
                                        
                                    indice_nber_glands = indice_nber_glands +1
                                    Variance_shape_glands,list1_closest_index = distance_center_cont (mask_ls)
                                    list_Variance_shape_glands.append(Variance_shape_glands)
                                    Final_image_lum =Final_image_lum +mask_ls
                                    Final_mask_full = Final_mask_full + mask_ls
                                   # cv2.imshow('',mask_ls)
                                   # cv2.waitKey(0)
                             if Final_image_mask_added_existence == True :
                                 ratio_density_stroma,ratio_density_nuclei,value_mac = distance_inter_lumen (img_nuclei_mask,Final_image_mask_added)         
                                 list_ratio_density_stroma.append(ratio_density_stroma)
                                 list_ratio_density_nuclei_1.append(ratio_density_nuclei)
                                 list_value_mac.append(value_mac)
           
            OUTPUT_DIR_name = OUTPUT_DIR + '/' + fname 
        #    cv2.imwrite(OUTPUT_DIR_name, Final_mask_full)
          #  print(np.shape(listresult_image1_))
           
            if Final_lumen_mask_added == False  and  indice_nber_glands>1:
                Variance_space_glands  = distance_inter_glands(Final_image_lum)
                list_space_glands.append(Variance_space_glands)
get_features_conf(1)   
#for indice_seg in range ()

               
    
